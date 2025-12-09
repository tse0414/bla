import os
from datetime import datetime
from zipfile import BadZipFile
import openpyxl
from openpyxl import Workbook

EXCEL_FILE = "logistics.xlsx"

# Parcels 欄位：狀態 → 金額 → 建立時間
PARCEL_HEADERS = [
    "TrackingNumber",   # 0
    "寄件人帳號",         # 1
    "收件人姓名",         # 2
    "收件地址",           # 3
    "重量",              # 4
    "服務類型",           # 5
    "狀態",              # 6
    "金額",              # 7
    "建立時間",           # 8
]

CUSTOMER_HEADERS = [
    "客戶帳號", "姓名", "電話", "Email", "地址", "建立時間"
]


# ------------------------------------------------
# 初始化 / 修復 Excel 結構
# ------------------------------------------------
EXCEL_FILE = "logistics.xlsx"

def initialize_excel():
    """檢查 Excel 檔，如果不存在就建立；如果壞掉就重建。"""

    def create_new_excel():
        from openpyxl import Workbook
        wb = Workbook()

        # Customers 工作表
        ws1 = wb.active
        ws1.title = "Customers"
        ws1.append([
            "客戶帳號", "姓名", "電話", "Email", "地址", "建立時間"
        ])

        # Parcels 工作表（含金額欄位）
        ws2 = wb.create_sheet("Parcels")
        ws2.append([
            "TrackingNumber",  # 0
            "寄件人帳號",        # 1
            "收件人姓名",        # 2
            "收件地址",          # 3
            "重量",             # 4
            "服務類型",          # 5
            "狀態",             # 6
            "金額",             # 7
            "建立時間",          # 8
        ])

        wb.save(EXCEL_FILE)

    # 檔案不存在 → 直接建立新的
    if not os.path.exists(EXCEL_FILE):
        create_new_excel()
        return

    # 檔案存在 → 檢查是不是合法的 xlsx
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        wb.close()
    except BadZipFile:
        # 檔案壞掉 → 先備份改名，再重建一份新的
        backup_name = f"broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{EXCEL_FILE}"
        os.rename(EXCEL_FILE, backup_name)
        create_new_excel()



# ------------------------------------------------
# Customers：新增 / 讀取 / 更新
# ------------------------------------------------
def append_customer(data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    ws.append([
        data.get("account"),
        data.get("name"),
        data.get("phone"),
        data.get("email"),
        data.get("address"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])

    wb.save(EXCEL_FILE)


def read_customers():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Customers" not in wb.sheetnames:
        return []

    ws = wb["Customers"]
    customers = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        customers.append({
            "account": row[0],
            "name": row[1],
            "phone": row[2],
            "email": row[3],
            "address": row[4],
            "created_at": row[5],
        })

    return customers


def update_customer(account, data):
    """
    以帳號為 key 更新姓名 / 電話 / Email / 地址，不動建立時間
    data 可有 name, phone, email, address
    """
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Customers" not in wb.sheetnames:
        return

    ws = wb["Customers"]

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(account):
            if "name" in data:
                row[1].value = data["name"]
            if "phone" in data:
                row[2].value = data["phone"]
            if "email" in data:
                row[3].value = data["email"]
            if "address" in data:
                row[4].value = data["address"]
            break

    wb.save(EXCEL_FILE)


# ------------------------------------------------
# Parcels：新增 / 讀取 / 更新金額
# ------------------------------------------------
def append_parcel(data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        ws = wb.create_sheet("Parcels")
        ws.append(PARCEL_HEADERS)
    else:
        ws = wb["Parcels"]

    ws.append([
        data.get("tracking_number"),
        data.get("sender_id"),
        data.get("recipient_name"),
        data.get("recipient_address"),
        data.get("weight"),
        data.get("service_type"),
        data.get("status"),
        data.get("amount"),
        data.get("created_at"),
    ])

    wb.save(EXCEL_FILE)


def read_parcels():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        return []

    ws = wb["Parcels"]
    parcels = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        parcels.append({
            "tracking_number": row[0],
            "sender_id": row[1],
            "recipient_name": row[2],
            "recipient_address": row[3],
            "weight": row[4],
            "service_type": row[5],
            "status": row[6],
            "amount": row[7],
            "created_at": row[8],
        })

    return parcels


def update_parcel_amount(tracking_number, amount):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        return

    ws = wb["Parcels"]

    for i, header in enumerate(PARCEL_HEADERS, start=1):
        ws.cell(row=1, column=i, value=header)

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(tracking_number):
            row[7].value = amount   # 金額欄
            break

    wb.save(EXCEL_FILE)
