import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from zipfile import BadZipFile

EXCEL_FILE = "logistics.xlsx"

CUSTOMER_HEADERS = [
    "客戶帳號",
    "姓名",
    "電話",
    "Email",
    "地址",
    "建立時間",
]

PARCEL_HEADERS = [
    "TrackingNumber",
    "寄件人帳號",
    "收件人姓名",
    "收件地址",
    "重量",
    "服務類型",
    "狀態",
    "金額",
    "建立時間",
]


# --------------------------------------------------
# 初始化 Excel
# --------------------------------------------------
def initialize_excel():

    def create_new():
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Customers"
        ws1.append(CUSTOMER_HEADERS)

        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)

        wb.save(EXCEL_FILE)

    if not os.path.exists(EXCEL_FILE):
        create_new()
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except BadZipFile:
        backup = f"broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{EXCEL_FILE}"
        os.rename(EXCEL_FILE, backup)
        create_new()
        return

    changed = False

    # Customers
    if "Customers" not in wb.sheetnames:
        ws = wb.create_sheet("Customers", 0)
        ws.append(CUSTOMER_HEADERS)
        changed = True
    else:
        ws = wb["Customers"]
        for i, h in enumerate(CUSTOMER_HEADERS, start=1):
            if ws.cell(row=1, column=i).value != h:
                ws.cell(row=1, column=i, value=h)
                changed = True

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=6).value in (None, ""):
                ws.cell(row=r, column=6, value=now)
                changed = True

    # Parcels
    if "Parcels" not in wb.sheetnames:
        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)
        changed = True

    if changed:
        wb.save(EXCEL_FILE)

    wb.close()


# --------------------------------------------------
# Customers
# --------------------------------------------------
def append_customer(data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        data.get("account", ""),
        data.get("name", ""),
        data.get("phone", ""),
        data.get("email", ""),
        data.get("address", ""),
        now,
    ])

    wb.save(EXCEL_FILE)
    wb.close()


def read_customers():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    result = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        result.append({
            "account": r[0],
            "name": r[1],
            "phone": r[2],
            "email": r[3],
            "address": r[4],
            "created_at": r[5],
        })

    wb.close()
    return result


def update_customer(account, data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == account:
            row[1].value = data.get("name", row[1].value)
            row[2].value = data.get("phone", row[2].value)
            row[3].value = data.get("email", row[3].value)
            row[4].value = data.get("address", row[4].value)
            break

    wb.save(EXCEL_FILE)
    wb.close()


# --------------------------------------------------
# Parcels（這就是你錯誤缺的部分）
# --------------------------------------------------
def append_parcel(record):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    ws.append([
        record.get("tracking_number"),
        record.get("sender_id"),
        record.get("recipient_name"),
        record.get("recipient_address"),
        record.get("weight"),
        record.get("service_type"),
        record.get("status"),
        record.get("amount"),
        record.get("created_at"),
    ])

    wb.save(EXCEL_FILE)
    wb.close()


def read_parcels():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    parcels = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        parcels.append({
            "tracking_number": r[0],
            "sender_id": r[1],
            "recipient_name": r[2],
            "recipient_address": r[3],
            "weight": r[4],
            "service_type": r[5],
            "status": r[6],
            "amount": r[7],
            "created_at": r[8],
        })

    wb.close()
    return parcels


def update_parcel_amount(tracking_number, amount):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == tracking_number:
            row[7].value = amount
            break

    wb.save(EXCEL_FILE)
    wb.close()


def update_parcel_status(tracking_number, status):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == tracking_number:
            row[6].value = status
            updated = True
            break

    wb.save(EXCEL_FILE)
    wb.close()
    return updated
