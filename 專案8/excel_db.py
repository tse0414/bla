import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from zipfile import BadZipFile

EXCEL_FILE = "logistics.xlsx"

CUSTOMER_HEADERS = [
    "客戶帳號",
    "姓名",
    "電話",
    "Email",
    "地址",
    "客戶類型",  
    "帳單偏好",  
    "建立時間",
]

PARCEL_HEADERS = [
    "TrackingNumber",
    "寄件人帳號",
    "收件人姓名",
    "收件地址",
    "重量",
    "包裹類型",   
    "申報價值",      
    "內容物描述",    
    "服務類型",
    "狀態",
    "金額",
    "建立時間",
    "付款狀態",
]


EVENT_HEADERS = [
    "事件ID",
    "追蹤編號",
    "事件類型",
    "時間戳記",
    "地點",
    "運輸載具ID",
    "倉儲地點",
    "操作人員",
    "備註",
]

# ✅ 新增: 帳號表 (解決帳號記憶體問題)
ACCOUNT_HEADERS = [
    "帳號",
    "密碼",  # 實際應用需加密
    "角色",
    "建立時間",
]


# --------------------------------------------------
# 初始化 Excel
# --------------------------------------------------
def initialize_excel():
    def create_new():
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Customers"
        ws1.append(CUSTOMER_HEADERS)

        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)
        
        ws3 = wb.create_sheet("TrackingEvents")
        ws3.append(EVENT_HEADERS)
        
        ws4 = wb.create_sheet("Accounts")
        ws4.append(ACCOUNT_HEADERS)

        wb.save(EXCEL_FILE)

    if not os.path.exists(EXCEL_FILE):
        create_new()
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except BadZipFile:
        backup = f"broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{EXCEL_FILE}"
        os.rename(EXCEL_FILE, backup)
        create_new()
        return

    changed = False

    # Customers
    if "Customers" not in wb.sheetnames:
        ws = wb.create_sheet("Customers", 0)
        ws.append(CUSTOMER_HEADERS)
        changed = True
    else:
        ws = wb["Customers"]
        # 檢查並補充欄位
        if ws.max_column < len(CUSTOMER_HEADERS):
            for i, h in enumerate(CUSTOMER_HEADERS, start=1):
                if ws.cell(row=1, column=i).value != h:
                    ws.cell(row=1, column=i, value=h)
                    changed = True

    # Parcels
    if "Parcels" not in wb.sheetnames:
        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)
        changed = True
    else:
        ws2 = wb["Parcels"]
        if ws2.max_column < len(PARCEL_HEADERS):
            for i, h in enumerate(PARCEL_HEADERS, start=1):
                if ws2.cell(row=1, column=i).value != h:
                    ws2.cell(row=1, column=i, value=h)
                    changed = True
    
    # ✅ TrackingEvents
    if "TrackingEvents" not in wb.sheetnames:
        ws3 = wb.create_sheet("TrackingEvents")
        ws3.append(EVENT_HEADERS)
        changed = True
    
    # ✅ Accounts
    if "Accounts" not in wb.sheetnames:
        ws4 = wb.create_sheet("Accounts")
        ws4.append(ACCOUNT_HEADERS)
        changed = True

    if changed:
        wb.save(EXCEL_FILE)

    wb.close()


# --------------------------------------------------
# 帳號管理 (解決記憶體問題)
# --------------------------------------------------
def append_account(account_data):
    """新增帳號到 Excel"""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Accounts"]
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        account_data.get("username", ""),
        account_data.get("password", ""),  # 實際應加密
        account_data.get("role", "customer"),
        now,
    ])
    
    wb.save(EXCEL_FILE)
    wb.close()


def read_accounts():
    """讀取所有帳號 (取代記憶體 ACCOUNTS)"""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Accounts"]
    
    accounts = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:  # 帳號不為空
            accounts[r[0]] = {
                "password": r[1],
                "role": r[2],
                "created_at": r[3],
            }
    
    wb.close()
    return accounts


def find_account(username):
    """查詢單一帳號"""
    accounts = read_accounts()
    return accounts.get(username)


# --------------------------------------------------
# Customers (修改版)
# --------------------------------------------------
def append_customer(data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        data.get("account", ""),
        data.get("name", ""),
        data.get("phone", ""),
        data.get("email", ""),
        data.get("address", ""),
        data.get("customer_type", "NON_CONTRACT"),  # ✅ 新增
        data.get("billing_preference", "COD"),      # ✅ 新增
        now,
    ])

    wb.save(EXCEL_FILE)
    wb.close()


def read_customers():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    result = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        result.append({
            "account": r[0],
            "name": r[1],
            "phone": r[2],
            "email": r[3],
            "address": r[4],
            "customer_type": r[5] if len(r) > 5 else "NON_CONTRACT",
            "billing_preference": r[6] if len(r) > 6 else "COD",
            "created_at": r[7] if len(r) > 7 else r[5],
        })

    wb.close()
    return result


def update_customer(account, data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == account:
            row[1].value = data.get("name", row[1].value)
            row[2].value = data.get("phone", row[2].value)
            row[3].value = data.get("email", row[3].value)
            row[4].value = data.get("address", row[4].value)
            row[5].value = data.get("customer_type", row[5].value)  # ✅
            row[6].value = data.get("billing_preference", row[6].value)  # ✅
            break

    wb.save(EXCEL_FILE)
    wb.close()


# --------------------------------------------------
# Parcels (修改版)
# --------------------------------------------------
def append_parcel(record):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    ws.append([
        record.get("tracking_number"),
        record.get("sender_id"),
        record.get("recipient_name"),
        record.get("recipient_address"),
        record.get("weight"),
        record.get("package_type", "中型箱"),      # ✅ 新增
        record.get("declared_value", 0),          # ✅ 新增
        record.get("contents", "一般貨物"),       # ✅ 新增
        record.get("service_type"),
        record.get("status"),
        record.get("amount"),
        record.get("created_at"),
        record.get("payment_status", "Unpaid"),
    ])

    wb.save(EXCEL_FILE)
    wb.close()


def read_parcels():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    parcels = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        parcels.append({
            "tracking_number": r[0],
            "sender_id": r[1],
            "recipient_name": r[2],
            "recipient_address": r[3],
            "weight": r[4],
            "package_type": r[5] if len(r) > 5 else "",
            "declared_value": r[6] if len(r) > 6 else 0,
            "contents": r[7] if len(r) > 7 else "",
            "service_type": r[8] if len(r) > 8 else r[5],
            "status": r[9] if len(r) > 9 else r[6],
            "amount": r[10] if len(r) > 10 else r[7],
            "created_at": r[11] if len(r) > 11 else r[8],
            "payment_status": r[12] if len(r) > 12 else "Unpaid",
        })

    wb.close()
    return parcels


def update_parcel_amount(tracking_number, amount):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == tracking_number:
            row[10].value = amount  # 調整索引
            break

    wb.save(EXCEL_FILE)
    wb.close()


def update_parcel_status(tracking_number, status):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Parcels"]

    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == tracking_number:
            row[9].value = status  # 調整索引
            updated = True
            break

    wb.save(EXCEL_FILE)
    wb.close()
    return updated


# --------------------------------------------------
# 物流事件追蹤
# --------------------------------------------------
def append_tracking_event(event):
    """記錄物流事件"""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["TrackingEvents"]
    
    ws.append([
        event.get("event_id"),
        event.get("tracking_number"),
        event.get("event_type"),
        event.get("timestamp"),
        event.get("location", ""),
        event.get("vehicle_id", ""),
        event.get("warehouse_id", ""),
        event.get("operator", ""),
        event.get("description", ""),
    ])
    
    wb.save(EXCEL_FILE)
    wb.close()


def read_tracking_events(tracking_number):
    """查詢包裹的完整追蹤歷史"""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["TrackingEvents"]
    
    events = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] == tracking_number:  # 追蹤編號匹配
            events.append({
                "event_id": r[0],
                "tracking_number": r[1],
                "event_type": r[2],
                "timestamp": r[3],
                "location": r[4],
                "vehicle_id": r[5],
                "warehouse_id": r[6],
                "operator": r[7],
                "description": r[8],
            })
    
    wb.close()
    # 按時間排序
    events.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return events


def read_all_tracking_events():
    """讀取所有事件 (用於報表分析)"""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["TrackingEvents"]
    
    events = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        events.append({
            "event_id": r[0],
            "tracking_number": r[1],
            "event_type": r[2],
            "timestamp": r[3],
            "location": r[4],
            "vehicle_id": r[5],
            "warehouse_id": r[6],
            "operator": r[7],
            "description": r[8],
        })
    
    wb.close()
    return events

def read_all_events_for_search():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["TrackingEvents"]
    
    events = []
    # 從第二行開始讀取 (第一行是標題)
    for r in ws.iter_rows(min_row=2, values_only=True):
        # r[1] 是追蹤編號, r[5] 是車輛ID, r[6] 是倉儲ID
        events.append({
            "tracking_number": r[1],
            "vehicle_id": r[5] if len(r) > 5 else "",
            "warehouse_id": r[6] if len(r) > 6 else "",
        })
    
    wb.close()
    return events