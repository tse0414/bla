import os
from datetime import datetime

import openpyxl
from openpyxl import Workbook

EXCEL_FILE = "logistics.xlsx"

# 你要的完整欄位（狀態 → 金額 → 建立時間）
PARCEL_HEADERS = [
    "TrackingNumber",   # 0
    "寄件人帳號",         # 1
    "收件人姓名",         # 2
    "收件地址",           # 3
    "重量",              # 4
    "服務類型",           # 5
    "狀態",              # 6
    "金額",              # 7
    "建立時間",           # 8
]


# ------------------------------------------------
# 初始化 / 修復 Excel 結構
# ------------------------------------------------
def initialize_excel():
    """
    確保：
      1. 檔案存在
      2. 一定有 Customers & Parcels 兩張表
      3. Parcels 的第 1 列欄位為 PARCEL_HEADERS
    """
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()

        # Customers
        ws1 = wb.active
        ws1.title = "Customers"
        ws1.append(["客戶帳號", "姓名", "電話", "Email", "地址", "建立時間"])

        # Parcels
        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)

        wb.save(EXCEL_FILE)
        return

    # 檔案已存在 → 檢查內容
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 確保 Customers 存在
    if "Customers" not in wb.sheetnames:
        ws1 = wb.create_sheet("Customers")
        ws1.append(["客戶帳號", "姓名", "電話", "Email", "地址", "建立時間"])

    # 確保 Parcels 存在＋標題正確
    if "Parcels" not in wb.sheetnames:
        ws2 = wb.create_sheet("Parcels")
        ws2.append(PARCEL_HEADERS)
    else:
        ws2 = wb["Parcels"]
        # 直接把第 1 列蓋成正確 header，避免舊檔欄位順序怪掉
        for idx, header in enumerate(PARCEL_HEADERS, start=1):
            ws2.cell(row=1, column=idx, value=header)

    wb.save(EXCEL_FILE)


# ------------------------------------------------
# 新增客戶（寫入 Customers）
# ------------------------------------------------
def append_customer(data):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Customers"]

    ws.append([
        data.get("account"),
        data.get("name"),
        data.get("phone"),
        data.get("email"),
        data.get("address"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])

    wb.save(EXCEL_FILE)


# ------------------------------------------------
# 新增包裹（寫入 Parcels：9 欄都寫）
# ------------------------------------------------
def append_parcel(data):
    """
    data 需要包含：
      tracking_number, sender_id, recipient_name, recipient_address,
      weight, service_type, status, created_at
    金額 amount 可以先是 None（付款後再補）
    """
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        ws = wb.create_sheet("Parcels")
        ws.append(PARCEL_HEADERS)
    else:
        ws = wb["Parcels"]

    ws.append([
        data.get("tracking_number"),    # 0
        data.get("sender_id"),          # 1
        data.get("recipient_name"),     # 2
        data.get("recipient_address"),  # 3
        data.get("weight"),             # 4
        data.get("service_type"),       # 5
        data.get("status"),             # 6
        data.get("amount"),             # 7 金額
        data.get("created_at"),         # 8 建立時間
    ])

    wb.save(EXCEL_FILE)


# ------------------------------------------------
# 讀取所有包裹（給 /records / 搜尋用）
# ------------------------------------------------
def read_parcels():
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        return []

    ws = wb["Parcels"]
    parcels = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        parcels.append({
            "tracking_number": row[0],
            "sender_id": row[1],
            "recipient_name": row[2],
            "recipient_address": row[3],
            "weight": row[4],
            "service_type": row[5],
            "status": row[6],
            "amount": row[7],
            "created_at": row[8],
        })

    return parcels


# ------------------------------------------------
# 更新指定追蹤編號的金額（付款後呼叫）
# ------------------------------------------------
def update_parcel_amount(tracking_number, amount):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Parcels" not in wb.sheetnames:
        return

    ws = wb["Parcels"]

    # 確保標題正確
    for idx, header in enumerate(PARCEL_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(tracking_number):
            # 金額在第 8 欄（索引 7）
            row[7].value = amount
            break

    wb.save(EXCEL_FILE)
